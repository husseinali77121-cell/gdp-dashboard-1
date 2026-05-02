"""
Biobase BK-310 Biochemistry Analyzer - Validation Program
Generates Excel templates for data entry and processes them to produce validation reports.
Modules: QC, Precision, Accuracy, Linearity, Carryover
"""

import pandas as pd
import numpy as np
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
import os
import sys
from datetime import datetime
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side, Protection
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.drawing.image import Image as XLImage

# ------------------------------
# Acceptance criteria for biochemistry tests
# Units and typical CLIA / desirable specs
# (You can edit these limits as you wish)
# Format: Test: (unit, precision_CV_max%, bias_max%, linearity_recovery_min%, linearity_recovery_max%, carryover_max%)
# ------------------------------
ACCEPTANCE = {
    "Glucose":         ("mg/dL", 3.0, 5.0, 90.0, 110.0, 1.0),
    "Urea":            ("mg/dL", 4.0, 7.0, 90.0, 110.0, 1.0),
    "Creatinine":      ("mg/dL", 4.0, 7.0, 90.0, 110.0, 1.0),
    "Uric Acid":       ("mg/dL", 5.0, 8.0, 90.0, 110.0, 1.0),
    "Total Protein":   ("g/dL", 3.0, 5.0, 90.0, 110.0, 1.0),
    "Albumin":         ("g/dL", 3.0, 5.0, 90.0, 110.0, 1.0),
    "Total Bilirubin": ("mg/dL", 5.0, 10.0, 90.0, 110.0, 1.0),
    "Direct Bilirubin":("mg/dL", 5.0, 10.0, 90.0, 110.0, 1.0),
    "ALT":             ("U/L", 5.0, 10.0, 90.0, 110.0, 1.0),
    "AST":             ("U/L", 5.0, 10.0, 90.0, 110.0, 1.0),
    "ALP":             ("U/L", 5.0, 10.0, 90.0, 110.0, 1.0),
    "GGT":             ("U/L", 5.0, 10.0, 90.0, 110.0, 1.0),
    "Cholesterol":     ("mg/dL", 3.0, 5.0, 90.0, 110.0, 1.0),
    "Triglycerides":   ("mg/dL", 5.0, 10.0, 90.0, 110.0, 1.0),
    "HDL":             ("mg/dL", 4.0, 10.0, 90.0, 110.0, 1.0),
    "LDL":             ("mg/dL", 4.0, 10.0, 90.0, 110.0, 1.0),
    "Calcium":         ("mg/dL", 2.0, 4.0, 90.0, 110.0, 1.0),
    "PO4":             ("mg/dL", 4.0, 8.0, 90.0, 110.0, 1.0),   # Phosphate
    "Magnesium":       ("mg/dL", 4.0, 8.0, 90.0, 110.0, 1.0),   # Mg
    "Zinc":            ("µg/dL", 5.0, 10.0, 90.0, 110.0, 1.0),
    "Iron":            ("µg/dL", 5.0, 10.0, 90.0, 110.0, 1.0),
    "Ferritin":        ("ng/mL", 8.0, 15.0, 90.0, 110.0, 1.0),
    "Vit D":           ("ng/mL", 8.0, 20.0, 90.0, 110.0, 1.0),
    "HbA1c":           ("%", 3.0, 5.0, 90.0, 110.0, 1.0),
}

OUTPUT_DIR = "output"
TEMPLATE_DIR = "templates"
os.makedirs(OUTPUT_DIR, exist_ok=True)
os.makedirs(TEMPLATE_DIR, exist_ok=True)

# =============== Styling helpers ===============
HEADER_FILL = PatternFill(start_color="D9EAD3", end_color="D9EAD3", fill_type="solid")
HEADER_FONT = Font(bold=True, size=11)
USER_INPUT_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
PROTECTED_FILL = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)
center_align = Alignment(horizontal='center', vertical='center')

def _style_header_row(ws, row, headers):
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = center_align
        cell.border = thin_border

def _apply_input_cell(ws, row, col, value=None):
    cell = ws.cell(row=row, column=col, value=value)
    cell.fill = USER_INPUT_FILL
    cell.border = thin_border
    cell.alignment = center_align
    return cell

def _apply_fixed_cell(ws, row, col, value, protected=False):
    cell = ws.cell(row=row, column=col, value=value)
    cell.fill = PROTECTED_FILL
    cell.border = thin_border
    cell.alignment = center_align
    if protected:
        cell.protection = Protection(locked=True)
    return cell

def _add_instruction_sheet(wb, module_name, instructions_text):
    ws = wb.create_sheet("Instructions")
    ws.append([f"Template: {module_name}"])
    ws.append(["Instructions:"])
    for line in instructions_text.split('\n'):
        ws.append([line])
    ws['A1'].font = Font(bold=True, size=12)
    return wb

# =============== Template Generators ===============
def generate_qc_template():
    wb = Workbook()
    ws = wb.active
    ws.title = "QC Data"
    headers = ["Test", "Level", "Lot", "Target_Mean", "Target_SD", "Measured_Value"]
    _style_header_row(ws, 1, headers)

    ws.merge_cells('A2:F2')
    note_cell = ws.cell(row=2, column=1, value="⬇️ Fill the YELLOW cells only. Test names and levels are pre-filled.")
    note_cell.font = Font(italic=True, color="555555")

    tests = list(ACCEPTANCE.keys())
    levels = ["Normal", "Patho"]
    row = 3
    for test in tests:
        for level in levels:
            _apply_fixed_cell(ws, row, 1, test, protected=True)
            _apply_fixed_cell(ws, row, 2, level, protected=True)
            _apply_input_cell(ws, row, 3)  # Lot
            _apply_input_cell(ws, row, 4)  # Target_Mean
            _apply_input_cell(ws, row, 5)  # Target_SD
            _apply_input_cell(ws, row, 6)  # Measured_Value
            row += 1

    ws.column_dimensions['A'].width = 22
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 14
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 18

    ws.protection.sheet = True
    ws.protection.set_password('')
    for r in range(3, row):
        for c in [3,4,5,6]:
            ws.cell(row=r, column=c).protection = Protection(locked=False)

    instructions = (
        "1. Fill Lot numbers, Target Mean, Target SD, and Measured Value.\n"
        "2. Yellow cells are unlocked; gray cells are locked.\n"
        "3. Save the file and upload for analysis."
    )
    _add_instruction_sheet(wb, "Quality Control", instructions)
    path = os.path.join(TEMPLATE_DIR, "qc_template.xlsx")
    wb.save(path)
    print(f"QC template saved to {path}")

def generate_precision_template(num_replicates=10):
    wb = Workbook()
    ws = wb.active
    ws.title = "Precision Data"
    headers = ["Test"] + [f"Rep_{i+1}" for i in range(num_replicates)]
    _style_header_row(ws, 1, headers)

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(headers))
    note_cell = ws.cell(row=2, column=1, value="⬇️ Fill the YELLOW cells with replicate results.")
    note_cell.font = Font(italic=True, color="555555")

    tests = list(ACCEPTANCE.keys())
    row = 3
    for test in tests:
        _apply_fixed_cell(ws, row, 1, test, protected=True)
        for rep_idx in range(num_replicates):
            _apply_input_cell(ws, row, rep_idx+2)
        row += 1

    ws.column_dimensions['A'].width = 22
    for i in range(num_replicates):
        col_letter = openpyxl.utils.get_column_letter(i+2)
        ws.column_dimensions[col_letter].width = 10

    ws.protection.sheet = True
    ws.protection.set_password('')
    for r in range(3, row):
        for c in range(2, num_replicates+2):
            ws.cell(row=r, column=c).protection = Protection(locked=False)

    instructions = (
        f"1. For each test, run the same sample {num_replicates} times on BK-310.\n"
        "2. Enter results in the yellow replicate columns.\n"
        "3. Save and upload."
    )
    _add_instruction_sheet(wb, "Precision", instructions)
    path = os.path.join(TEMPLATE_DIR, "precision_template.xlsx")
    wb.save(path)
    print(f"Precision template saved to {path}")

def generate_accuracy_template():
    wb = Workbook()
    ws = wb.active
    ws.title = "Accuracy Data"
    headers = ["Test", "Reference_Value", "Measured_Value"]
    _style_header_row(ws, 1, headers)

    ws.merge_cells('A2:C2')
    note_cell = ws.cell(row=2, column=1, value="⬇️ Fill YELLOW cells with reference and measured values.")
    note_cell.font = Font(italic=True, color="555555")

    tests = list(ACCEPTANCE.keys())
    row = 3
    for test in tests:
        _apply_fixed_cell(ws, row, 1, test, protected=True)
        _apply_input_cell(ws, row, 2)  # Reference
        _apply_input_cell(ws, row, 3)  # Measured
        row += 1

    ws.column_dimensions['A'].width = 22
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 18

    ws.protection.sheet = True
    ws.protection.set_password('')
    for r in range(3, row):
        for c in [2,3]:
            ws.cell(row=r, column=c).protection = Protection(locked=False)

    instructions = (
        "1. Enter the known reference value (calibrator or reference method).\n"
        "2. Run the same sample on BK-310 and enter the measured value.\n"
        "3. Save and upload."
    )
    _add_instruction_sheet(wb, "Accuracy", instructions)
    path = os.path.join(TEMPLATE_DIR, "accuracy_template.xlsx")
    wb.save(path)
    print(f"Accuracy template saved to {path}")

def generate_linearity_template():
    wb = Workbook()
    ws = wb.active
    ws.title = "Linearity Data"
    headers = ["Test", "Expected_Conc", "Measured_Conc"]
    _style_header_row(ws, 1, headers)

    ws.merge_cells('A2:C2')
    note_cell = ws.cell(row=2, column=1, value="⬇️ Fill Expected_Conc and Measured_Conc (yellow). Use at least 5 dilutions per test.")
    note_cell.font = Font(italic=True, color="555555")

    tests = list(ACCEPTANCE.keys())
    row = 3
    for test in tests:
        # Create 5 dilution rows for each test
        for _ in range(5):
            _apply_fixed_cell(ws, row, 1, test, protected=True)
            _apply_input_cell(ws, row, 2)  # Expected_Conc
            _apply_input_cell(ws, row, 3)  # Measured_Conc
            row += 1

    ws.column_dimensions['A'].width = 22
    ws.column_dimensions['B'].width = 16
    ws.column_dimensions['C'].width = 16

    ws.protection.sheet = True
    ws.protection.set_password('')
    for r in range(3, row):
        for c in [2,3]:
            ws.cell(row=r, column=c).protection = Protection(locked=False)

    instructions = (
        "1. Prepare a series of dilutions (e.g., 0%, 25%, 50%, 75%, 100% of high pool).\n"
        "2. Write the calculated Expected_Conc for each dilution.\n"
        "3. Run each dilution on BK-310 and record Measured_Conc.\n"
        "4. The report will compute linear regression and recovery."
    )
    _add_instruction_sheet(wb, "Linearity", instructions)
    path = os.path.join(TEMPLATE_DIR, "linearity_template.xlsx")
    wb.save(path)
    print(f"Linearity template saved to {path}")

def generate_carryover_template():
    wb = Workbook()
    ws = wb.active
    ws.title = "Carryover Data"
    headers = ["Test", "High1", "High2", "High3", "Low1", "Low2", "Low3"]
    _style_header_row(ws, 1, headers)

    ws.merge_cells('A2:G2')
    note_cell = ws.cell(row=2, column=1, value="⬇️ Fill all yellow cells with the results from carryover experiment.")
    note_cell.font = Font(italic=True, color="555555")

    tests = list(ACCEPTANCE.keys())
    row = 3
    for test in tests:
        _apply_fixed_cell(ws, row, 1, test, protected=True)
        for col in range(2, 8):  # High1 to Low3
            _apply_input_cell(ws, row, col)
        row += 1

    ws.column_dimensions['A'].width = 22
    for col_letter in ['B','C','D','E','F','G']:
        ws.column_dimensions[col_letter].width = 10

    ws.protection.sheet = True
    ws.protection.set_password('')
    for r in range(3, row):
        for c in range(2, 8):
            ws.cell(row=r, column=c).protection = Protection(locked=False)

    instructions = (
        "1. Run a high-concentration sample three times consecutively (High1, High2, High3).\n"
        "2. Immediately run a low-concentration sample three times (Low1, Low2, Low3).\n"
        "3. Enter results. Carryover% is calculated automatically."
    )
    _add_instruction_sheet(wb, "Carryover", instructions)
    path = os.path.join(TEMPLATE_DIR, "carryover_template.xlsx")
    wb.save(path)
    print(f"Carryover template saved to {path}")

# ------------------------------
# Statistical helpers
# ------------------------------
def calc_stats(series):
    vals = series.dropna().astype(float)
    if len(vals) < 2:
        return np.nan, np.nan, np.nan
    mean = vals.mean()
    sd = vals.std(ddof=1)
    cv = (sd / mean * 100) if mean != 0 else np.nan
    return mean, sd, cv

def status(pass_condition):
    return "PASS" if pass_condition else "FAIL"

def plot_to_excel(ws, img_path, anchor):
    img = XLImage(img_path)
    img.anchor = anchor
    ws.add_image(img)

# ------------------------------
# Processing functions (unchanged logic, just read sheet name)
# ------------------------------
def process_qc(filename):
    df = pd.read_excel(filename, sheet_name="QC Data")
    required_cols = {"Test", "Level", "Target_Mean", "Target_SD", "Measured_Value"}
    if not required_cols.issubset(df.columns):
        raise ValueError(f"Missing columns. Required: {required_cols}")
    df["Target_Mean"] = pd.to_numeric(df["Target_Mean"], errors='coerce')
    df["Target_SD"] = pd.to_numeric(df["Target_SD"], errors='coerce')
    df["Measured_Value"] = pd.to_numeric(df["Measured_Value"], errors='coerce')
    df = df.dropna(subset=["Measured_Value"])
    df["Z_Score"] = (df["Measured_Value"] - df["Target_Mean"]) / df["Target_SD"]
    df["Acceptable"] = df["Z_Score"].abs() <= 2
    df["Status"] = df["Acceptable"].apply(lambda x: status(x))

    summary = df[["Test", "Level", "Lot", "Target_Mean", "Target_SD", "Measured_Value", "Z_Score", "Status"]]

    plot_files = []
    for (test, level), group in df.groupby(["Test", "Level"]):
        plt.figure()
        runs = range(1, len(group)+1)
        means = group["Target_Mean"].values
        sds = group["Target_SD"].values
        vals = group["Measured_Value"].values
        for i, (m, sd) in enumerate(zip(means, sds)):
            plt.axhline(y=m+2*sd, color='red', linestyle='--', linewidth=0.8)
            plt.axhline(y=m-2*sd, color='red', linestyle='--', linewidth=0.8)
            plt.axhline(y=m, color='green', linestyle='-', linewidth=0.8)
        plt.plot(runs, vals, marker='o', color='blue')
        plt.title(f"Levey-Jennings: {test} ({level})")
        plt.xlabel("Run")
        plt.ylabel("Result")
        plt.ylim(min(means - 3*sds) - 1, max(means + 3*sds) + 1)
        fname = os.path.join(OUTPUT_DIR, f"qc_lj_{test}_{level}.png")
        plt.savefig(fname)
        plt.close()
        plot_files.append(fname)

    wb = Workbook()
    ws = wb.active
    ws.title = "QC Results"
    for r in dataframe_to_rows(summary, index=False, header=True):
        ws.append(r)
    img_row = len(summary) + 3
    for i, fname in enumerate(plot_files):
        plot_to_excel(ws, fname, f"A{img_row}")
        img_row += 20
    out_path = os.path.join(OUTPUT_DIR, "QC_Report.xlsx")
    wb.save(out_path)
    print(f"QC report saved to {out_path}")
    return summary

def process_precision(filename):
    df = pd.read_excel(filename, sheet_name="Precision Data")
    test_col = "Test"
    rep_cols = [c for c in df.columns if c.startswith("Rep_")]
    melted = df.melt(id_vars=[test_col], value_vars=rep_cols, var_name="Replicate", value_name="Result")
    melted["Result"] = pd.to_numeric(melted["Result"], errors='coerce')
    grouped = melted.groupby("Test")["Result"]
    stats = grouped.agg(['mean', 'std', 'count'])
    stats['CV%'] = (stats['std'] / stats['mean']) * 100
    stats = stats.rename(columns={'mean':'Mean','std':'SD','count':'N'})
    stats = stats.reset_index()

    results = []
    for _, row in stats.iterrows():
        test = row["Test"]
        cv = row["CV%"]
        if test in ACCEPTANCE:
            limit = ACCEPTANCE[test][1]
            unit = ACCEPTANCE[test][0]
            pass_fail = status(cv <= limit)
        else:
            limit = None
            unit = "?"
            pass_fail = "Unknown"
        results.append({
            "Test": test,
            "Unit": unit,
            "Mean": round(row["Mean"], 2),
            "SD": round(row["SD"], 2),
            "CV%": round(cv, 2),
            "Acceptable CV%": limit,
            "Status": pass_fail
        })
    summary = pd.DataFrame(results)

    plt.figure()
    tests = summary["Test"]
    cv_vals = summary["CV%"]
    limits = summary["Acceptable CV%"]
    idx = np.arange(len(tests))
    width = 0.35
    plt.bar(idx, cv_vals, width, label='Measured CV%', color='skyblue')
    plt.bar(idx + width, limits, width, label='Max Allowed CV%', color='orange')
    plt.xticks(idx + width/2, tests, rotation=45)
    plt.ylabel("CV%")
    plt.title("Precision: CV% vs Limit")
    plt.legend()
    plot_path = os.path.join(OUTPUT_DIR, "precision_cv_chart.png")
    plt.tight_layout()
    plt.savefig(plot_path)
    plt.close()

    wb = Workbook()
    ws = wb.active
    ws.title = "Precision Results"
    for r in dataframe_to_rows(summary, index=False, header=True):
        ws.append(r)
    plot_to_excel(ws, plot_path, f"A{len(summary)+3}")
    out_path = os.path.join(OUTPUT_DIR, "Precision_Report.xlsx")
    wb.save(out_path)
    print(f"Precision report saved to {out_path}")
    return summary

def process_accuracy(filename):
    df = pd.read_excel(filename, sheet_name="Accuracy Data")
    df["Reference_Value"] = pd.to_numeric(df["Reference_Value"], errors='coerce')
    df["Measured_Value"] = pd.to_numeric(df["Measured_Value"], errors='coerce')
    df.dropna(subset=["Reference_Value", "Measured_Value"], inplace=True)
    df["Bias%"] = ((df["Measured_Value"] - df["Reference_Value"]) / df["Reference_Value"]) * 100

    results = []
    for _, row in df.iterrows():
        test = row["Test"]
        bias = row["Bias%"]
        if test in ACCEPTANCE:
            limit = ACCEPTANCE[test][2]
            unit = ACCEPTANCE[test][0]
            pass_fail = status(abs(bias) <= limit)
        else:
            limit = None
            unit = "?"
            pass_fail = "Unknown"
        results.append({
            "Test": test,
            "Unit": unit,
            "Reference": row["Reference_Value"],
            "Measured": row["Measured_Value"],
            "Bias%": round(bias, 2),
            "Acceptable Bias%": limit,
            "Status": pass_fail
        })
    summary = pd.DataFrame(results)

    plt.figure()
    for test in summary["Test"].unique():
        subset = summary[summary["Test"] == test]
        plt.scatter(subset["Reference"], subset["Measured"], label=test)
    plt.plot([summary["Reference"].min(), summary["Reference"].max()],
             [summary["Reference"].min(), summary["Reference"].max()], 'k--')
    plt.xlabel("Reference Value")
    plt.ylabel("BK-310 Measured")
    plt.title("Accuracy: Measured vs Reference")
    plt.legend()
    plot_path = os.path.join(OUTPUT_DIR, "accuracy_scatter.png")
    plt.savefig(plot_path)
    plt.close()

    wb = Workbook()
    ws = wb.active
    ws.title = "Accuracy Results"
    for r in dataframe_to_rows(summary, index=False, header=True):
        ws.append(r)
    plot_to_excel(ws, plot_path, f"A{len(summary)+3}")
    out_path = os.path.join(OUTPUT_DIR, "Accuracy_Report.xlsx")
    wb.save(out_path)
    print(f"Accuracy report saved to {out_path}")
    return summary

def process_linearity(filename):
    df = pd.read_excel(filename, sheet_name="Linearity Data")
    df["Expected_Conc"] = pd.to_numeric(df["Expected_Conc"], errors='coerce')
    df["Measured_Conc"] = pd.to_numeric(df["Measured_Conc"], errors='coerce')
    df.dropna(inplace=True)

    results = []
    plot_files = []
    for test, group in df.groupby("Test"):
        group = group.sort_values("Expected_Conc")
        x = group["Expected_Conc"].values
        y = group["Measured_Conc"].values
        if len(x) < 3:
            continue
        coeffs = np.polyfit(x, y, 1)
        slope, intercept = coeffs
        y_pred = slope * x + intercept
        recovery = (y / x) * 100
        mask = x > 0
        avg_recovery = np.mean(recovery[mask])
        if test in ACCEPTANCE:
            low_rec, high_rec = ACCEPTANCE[test][3], ACCEPTANCE[test][4]
            pass_fail = status(low_rec <= avg_recovery <= high_rec)
            unit = ACCEPTANCE[test][0]
        else:
            low_rec, high_rec = None, None
            pass_fail = "Unknown"
            unit = "?"
        results.append({
            "Test": test,
            "Unit": unit,
            "Slope": round(slope, 3),
            "Intercept": round(intercept, 3),
            "Avg Recovery%": round(avg_recovery, 2) if not np.isnan(avg_recovery) else "N/A",
            "Recovery Range": f"{low_rec}-{high_rec}%" if low_rec else "N/A",
            "Status": pass_fail
        })

        plt.figure()
        plt.scatter(x, y, label="Measured")
        plt.plot(x, y_pred, 'r-', label=f"Fit: y={slope:.3f}x+{intercept:.3f}")
        plt.xlabel("Expected Concentration")
        plt.ylabel("Measured Concentration")
        plt.title(f"Linearity: {test}")
        plt.legend()
        fname = os.path.join(OUTPUT_DIR, f"linearity_{test}.png")
        plt.savefig(fname)
        plt.close()
        plot_files.append(fname)

    summary = pd.DataFrame(results)
    wb = Workbook()
    ws = wb.active
    ws.title = "Linearity Results"
    for r in dataframe_to_rows(summary, index=False, header=True):
        ws.append(r)
    img_row = len(summary) + 3
    for f in plot_files:
        plot_to_excel(ws, f, f"A{img_row}")
        img_row += 20
    out_path = os.path.join(OUTPUT_DIR, "Linearity_Report.xlsx")
    wb.save(out_path)
    print(f"Linearity report saved to {out_path}")
    return summary

def process_carryover(filename):
    df = pd.read_excel(filename, sheet_name="Carryover Data")
    high_cols = ["High1", "High2", "High3"]
    low_cols = ["Low1", "Low2", "Low3"]
    results = []
    for _, row in df.iterrows():
        test = row["Test"]
        highs = pd.to_numeric(row[high_cols], errors='coerce').dropna()
        lows = pd.to_numeric(row[low_cols], errors='coerce').dropna()
        if len(highs) < 2 or len(lows) < 2:
            continue
        if len(lows) >= 3:
            low_first = lows.iloc[0]
            low_rest_mean = lows.iloc[1:].mean()
            high_all_mean = highs.mean()
            carry = ((low_first - low_rest_mean) / (high_all_mean - low_rest_mean)) * 100 if high_all_mean != low_rest_mean else np.nan
        else:
            carry = np.nan
        if test in ACCEPTANCE:
            limit = ACCEPTANCE[test][5]
            pass_fail = status(abs(carry) <= limit) if not np.isnan(carry) else "Invalid"
            unit = ACCEPTANCE[test][0]
        else:
            limit = None
            unit = "?"
            pass_fail = "Unknown"
        results.append({
            "Test": test,
            "Unit": unit,
            "Carryover%": round(carry, 2) if not np.isnan(carry) else "N/A",
            "Max Allowed%": limit,
            "Status": pass_fail
        })
    summary = pd.DataFrame(results)
    wb = Workbook()
    ws = wb.active
    ws.title = "Carryover Results"
    for r in dataframe_to_rows(summary, index=False, header=True):
        ws.append(r)
    out_path = os.path.join(OUTPUT_DIR, "Carryover_Report.xlsx")
    wb.save(out_path)
    print(f"Carryover report saved to {out_path}")
    return summary

if __name__ == "__main__":
    import argparse
    parser = argparse.ArgumentParser(description="BK-310 Biochemistry Validation")
    parser.add_argument("--module", type=str, required=True,
                        choices=["qc", "precision", "accuracy", "linearity", "carryover"])
    parser.add_argument("--generate", action="store_true")
    parser.add_argument("--process", type=str)
    parser.add_argument("--replicates", type=int, default=10)
    args = parser.parse_args()
    if args.module == "qc":
        if args.generate: generate_qc_template()
        elif args.process: process_qc(args.process)
    elif args.module == "precision":
        if args.generate: generate_precision_template(args.replicates)
        elif args.process: process_precision(args.process)
    elif args.module == "accuracy":
        if args.generate: generate_accuracy_template()
        elif args.process: process_accuracy(args.process)
    elif args.module == "linearity":
        if args.generate: generate_linearity_template()
        elif args.process: process_linearity(args.process)
    elif args.module == "carryover":
        if args.generate: generate_carryover_template()
        elif args.process: process_carryover(args.process)
