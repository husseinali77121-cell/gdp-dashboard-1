"""
Biobase BK-310 Biochemistry Analyzer - Validation Program
Generates Excel templates for data entry and processes them to produce validation reports.
Modules: QC, Precision, Accuracy, Linearity, Carryover
"""

import pandas as pd
import numpy as np
import matplotlib
matplotlib.use('Agg')  # non-interactive backend
import matplotlib.pyplot as plt
import os
import sys
from datetime import datetime
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.drawing.image import Image as XLImage
import argparse

# ------------------------------
# Configuration: acceptance criteria for biochemistry tests
# Units and typical CLIA / desirable specs
# ------------------------------
ACCEPTANCE = {
    # Test: (unit, precision_CV_max%, bias_max%, linearity_recovery_min%, linearity_recovery_max%, carryover_max%)
    "Glucose":    ("mg/dL", 3.0, 5.0, 90.0, 110.0, 1.0),
    "Urea":       ("mg/dL", 4.0, 7.0, 90.0, 110.0, 1.0),
    "Creatinine": ("mg/dL", 4.0, 7.0, 90.0, 110.0, 1.0),
    "Uric Acid":  ("mg/dL", 5.0, 8.0, 90.0, 110.0, 1.0),
    "Total Protein": ("g/dL", 3.0, 5.0, 90.0, 110.0, 1.0),
    "Albumin":    ("g/dL", 3.0, 5.0, 90.0, 110.0, 1.0),
    "Total Bilirubin": ("mg/dL", 5.0, 10.0, 90.0, 110.0, 1.0),
    "Direct Bilirubin": ("mg/dL", 5.0, 10.0, 90.0, 110.0, 1.0),
    "ALT":        ("U/L", 5.0, 10.0, 90.0, 110.0, 1.0),
    "AST":        ("U/L", 5.0, 10.0, 90.0, 110.0, 1.0),
    "ALP":        ("U/L", 5.0, 10.0, 90.0, 110.0, 1.0),
    "GGT":        ("U/L", 5.0, 10.0, 90.0, 110.0, 1.0),
    "Cholesterol": ("mg/dL", 3.0, 5.0, 90.0, 110.0, 1.0),
    "Triglycerides": ("mg/dL", 5.0, 10.0, 90.0, 110.0, 1.0),
    "HDL":        ("mg/dL", 4.0, 10.0, 90.0, 110.0, 1.0),
    "LDL":        ("mg/dL", 4.0, 10.0, 90.0, 110.0, 1.0),
}
# You can add more analytes or adjust limits as needed.

OUTPUT_DIR = "output"
TEMPLATE_DIR = "templates"
os.makedirs(OUTPUT_DIR, exist_ok=True)
os.makedirs(TEMPLATE_DIR, exist_ok=True)

# ------------------------------
# Helper functions
# ------------------------------
def calc_stats(series):
    """Return mean, SD, CV% for a numeric series."""
    vals = series.dropna().astype(float)
    if len(vals) < 2:
        return np.nan, np.nan, np.nan
    mean = vals.mean()
    sd = vals.std(ddof=1)  # sample SD
    cv = (sd / mean * 100) if mean != 0 else np.nan
    return mean, sd, cv

def status(pass_condition):
    return "PASS" if pass_condition else "FAIL"

def plot_to_excel(ws, img_path, anchor):
    """Insert a PNG image into an openpyxl worksheet."""
    img = XLImage(img_path)
    img.anchor = anchor
    ws.add_image(img)

# ------------------------------
# 1. QUALITY CONTROL MODULE
# ------------------------------
def generate_qc_template():
    """Create QC template: columns for Test, Level, Lot, Target Mean, Target SD, Measured Value."""
    template = pd.DataFrame(columns=["Test", "Level", "Lot", "Target_Mean", "Target_SD", "Measured_Value"])
    # Add example rows
    examples = [
        {"Test": "Glucose", "Level": "Normal", "Lot": "QC123", "Target_Mean": 100, "Target_SD": 3, "Measured_Value": ""},
        {"Test": "Glucose", "Level": "Patho", "Lot": "QC456", "Target_Mean": 250, "Target_SD": 7.5, "Measured_Value": ""},
    ]
    template = pd.concat([template, pd.DataFrame(examples)], ignore_index=True)
    path = os.path.join(TEMPLATE_DIR, "qc_template.xlsx")
    template.to_excel(path, index=False)
    print(f"QC template saved to {path}\nPlease fill in the 'Measured_Value' column and re-load.")

def process_qc(filename):
    """Read filled QC template, compute z-scores and Levey-Jennings plot per test/level."""
    df = pd.read_excel(filename)
    required_cols = {"Test", "Level", "Target_Mean", "Target_SD", "Measured_Value"}
    if not required_cols.issubset(df.columns):
        raise ValueError(f"Missing columns. Required: {required_cols}")
    df["Target_Mean"] = pd.to_numeric(df["Target_Mean"], errors='coerce')
    df["Target_SD"] = pd.to_numeric(df["Target_SD"], errors='coerce')
    df["Measured_Value"] = pd.to_numeric(df["Measured_Value"], errors='coerce')
    df = df.dropna(subset=["Measured_Value"])
    df["Z_Score"] = (df["Measured_Value"] - df["Target_Mean"]) / df["Target_SD"]
    df["Acceptable"] = df["Z_Score"].abs() <= 2  # within 2 SD
    df["Status"] = df["Acceptable"].apply(lambda x: status(x))

    # Summary per test & level
    summary = df[["Test", "Level", "Lot", "Target_Mean", "Target_SD", "Measured_Value", "Z_Score", "Status"]]

    # Generate Levey-Jennings plot per test-level combination
    plot_files = []
    for (test, level), group in df.groupby(["Test", "Level"]):
        plt.figure()
        runs = range(1, len(group)+1)
        means = group["Target_Mean"].values
        sds = group["Target_SD"].values
        vals = group["Measured_Value"].values
        z = group["Z_Score"].values
        # Plot +2SD, -2SD, mean lines
        for i, (m, sd) in enumerate(zip(means, sds)):
            plt.axhline(y=m+2*sd, color='red', linestyle='--', linewidth=0.8)
            plt.axhline(y=m-2*sd, color='red', linestyle='--', linewidth=0.8)
            plt.axhline(y=m, color='green', linestyle='-', linewidth=0.8)
        plt.plot(runs, vals, marker='o', color='blue')
        plt.title(f"Levey-Jennings: {test} ({level})")
        plt.xlabel("Run")
        plt.ylabel("Result")
        plt.ylim(min(means - 3*sds) - 1, max(means + 3*sds) + 1)
        fname = os.path.join(OUTPUT_DIR, f"qc_lj_{test}_{level}.png")
        plt.savefig(fname)
        plt.close()
        plot_files.append(fname)

    # Write Excel report
    wb = Workbook()
    ws = wb.active
    ws.title = "QC Results"
    for r in dataframe_to_rows(summary, index=False, header=True):
        ws.append(r)
    # Insert plots into same workbook
    img_row = len(summary) + 3
    for i, fname in enumerate(plot_files):
        plot_to_excel(ws, fname, f"A{img_row}")
        img_row += 20  # adjust spacing
    out_path = os.path.join(OUTPUT_DIR, "QC_Report.xlsx")
    wb.save(out_path)
    print(f"QC report saved to {out_path}")
    return summary

# ------------------------------
# 2. PRECISION (Within-Run) MODULE
# ------------------------------
def generate_precision_template(num_replicates=10):
    """Template: rows per test, replicate columns."""
    cols = ["Test"] + [f"Rep_{i+1}" for i in range(num_replicates)]
    template = pd.DataFrame(columns=cols)
    # Add one example row
    example = {"Test": "Glucose"}
    for i in range(num_replicates):
        example[f"Rep_{i+1}"] = ""
    template = pd.concat([template, pd.DataFrame([example])], ignore_index=True)
    path = os.path.join(TEMPLATE_DIR, "precision_template.xlsx")
    template.to_excel(path, index=False)
    print(f"Precision template saved to {path} (with {num_replicates} replicates). Fill and reload.")

def process_precision(filename):
    """Calculate mean, SD, CV% per test and compare to limit."""
    df = pd.read_excel(filename)
    test_col = "Test"
    rep_cols = [c for c in df.columns if c.startswith("Rep_")]
    # Melt to long format for easier stats
    melted = df.melt(id_vars=[test_col], value_vars=rep_cols, var_name="Replicate", value_name="Result")
    melted["Result"] = pd.to_numeric(melted["Result"], errors='coerce')
    grouped = melted.groupby("Test")["Result"]
    stats = grouped.agg(['mean', 'std', 'count'])
    stats['CV%'] = (stats['std'] / stats['mean']) * 100
    stats = stats.rename(columns={'mean':'Mean','std':'SD','count':'N'})
    stats = stats.reset_index()

    results = []
    for _, row in stats.iterrows():
        test = row["Test"]
        cv = row["CV%"]
        if test in ACCEPTANCE:
            limit = ACCEPTANCE[test][1]  # precision CV% limit
            unit = ACCEPTANCE[test][0]
            pass_fail = status(cv <= limit)
        else:
            limit = None
            unit = "?"
            pass_fail = "Unknown"
        results.append({
            "Test": test,
            "Unit": unit,
            "Mean": round(row["Mean"], 2),
            "SD": round(row["SD"], 2),
            "CV%": round(cv, 2),
            "Acceptable CV%": limit,
            "Status": pass_fail
        })
    summary = pd.DataFrame(results)

    # Plot CV% bar chart vs limits
    plt.figure()
    tests = summary["Test"]
    cv_vals = summary["CV%"]
    limits = summary["Acceptable CV%"]
    idx = np.arange(len(tests))
    width = 0.35
    plt.bar(idx, cv_vals, width, label='Measured CV%', color='skyblue')
    plt.bar(idx + width, limits, width, label='Max Allowed CV%', color='orange')
    plt.xticks(idx + width/2, tests, rotation=45)
    plt.ylabel("CV%")
    plt.title("Precision: CV% vs Limit")
    plt.legend()
    plot_path = os.path.join(OUTPUT_DIR, "precision_cv_chart.png")
    plt.tight_layout()
    plt.savefig(plot_path)
    plt.close()

    # Excel report
    wb = Workbook()
    ws = wb.active
    ws.title = "Precision Results"
    for r in dataframe_to_rows(summary, index=False, header=True):
        ws.append(r)
    plot_to_excel(ws, plot_path, f"A{len(summary)+3}")
    out_path = os.path.join(OUTPUT_DIR, "Precision_Report.xlsx")
    wb.save(out_path)
    print(f"Precision report saved to {out_path}")
    return summary

# ------------------------------
# 3. ACCURACY (Bias) MODULE
# ------------------------------
def generate_accuracy_template():
    template = pd.DataFrame(columns=["Test", "Reference_Value", "Measured_Value"])
    examples = [
        {"Test": "Glucose", "Reference_Value": "", "Measured_Value": ""},
    ]
    template = pd.concat([template, pd.DataFrame(examples)], ignore_index=True)
    path = os.path.join(TEMPLATE_DIR, "accuracy_template.xlsx")
    template.to_excel(path, index=False)
    print(f"Accuracy template saved to {path}")

def process_accuracy(filename):
    df = pd.read_excel(filename)
    df["Reference_Value"] = pd.to_numeric(df["Reference_Value"], errors='coerce')
    df["Measured_Value"] = pd.to_numeric(df["Measured_Value"], errors='coerce')
    df.dropna(subset=["Reference_Value", "Measured_Value"], inplace=True)
    df["Bias%"] = ((df["Measured_Value"] - df["Reference_Value"]) / df["Reference_Value"]) * 100

    results = []
    for _, row in df.iterrows():
        test = row["Test"]
        bias = row["Bias%"]
        if test in ACCEPTANCE:
            limit = ACCEPTANCE[test][2]  # bias% limit
            unit = ACCEPTANCE[test][0]
            pass_fail = status(abs(bias) <= limit)
        else:
            limit = None
            unit = "?"
            pass_fail = "Unknown"
        results.append({
            "Test": test,
            "Unit": unit,
            "Reference": row["Reference_Value"],
            "Measured": row["Measured_Value"],
            "Bias%": round(bias, 2),
            "Acceptable Bias%": limit,
            "Status": pass_fail
        })
    summary = pd.DataFrame(results)

    # Scatter plot measured vs reference
    plt.figure()
    for test in summary["Test"].unique():
        subset = summary[summary["Test"] == test]
        plt.scatter(subset["Reference"], subset["Measured"], label=test)
    plt.plot([summary["Reference"].min(), summary["Reference"].max()],
             [summary["Reference"].min(), summary["Reference"].max()], 'k--')
    plt.xlabel("Reference Value")
    plt.ylabel("BK-310 Measured")
    plt.title("Accuracy: Measured vs Reference")
    plt.legend()
    plot_path = os.path.join(OUTPUT_DIR, "accuracy_scatter.png")
    plt.savefig(plot_path)
    plt.close()

    wb = Workbook()
    ws = wb.active
    ws.title = "Accuracy Results"
    for r in dataframe_to_rows(summary, index=False, header=True):
        ws.append(r)
    plot_to_excel(ws, plot_path, f"A{len(summary)+3}")
    out_path = os.path.join(OUTPUT_DIR, "Accuracy_Report.xlsx")
    wb.save(out_path)
    print(f"Accuracy report saved to {out_path}")
    return summary

# ------------------------------
# 4. LINEARITY MODULE
# ------------------------------
def generate_linearity_template():
    """Template: Test, Dilution_Factor (or Expected_Conc), Measured_Conc"""
    template = pd.DataFrame(columns=["Test", "Expected_Conc", "Measured_Conc"])
    # Example for Glucose
    example = [
        {"Test": "Glucose", "Expected_Conc": 0, "Measured_Conc": ""},
        {"Test": "Glucose", "Expected_Conc": 50, "Measured_Conc": ""},
        {"Test": "Glucose", "Expected_Conc": 100, "Measured_Conc": ""},
        {"Test": "Glucose", "Expected_Conc": 200, "Measured_Conc": ""},
        {"Test": "Glucose", "Expected_Conc": 400, "Measured_Conc": ""},
    ]
    template = pd.concat([template, pd.DataFrame(example)], ignore_index=True)
    path = os.path.join(TEMPLATE_DIR, "linearity_template.xlsx")
    template.to_excel(path, index=False)
    print(f"Linearity template saved to {path}")

def process_linearity(filename):
    df = pd.read_excel(filename)
    df["Expected_Conc"] = pd.to_numeric(df["Expected_Conc"], errors='coerce')
    df["Measured_Conc"] = pd.to_numeric(df["Measured_Conc"], errors='coerce')
    df.dropna(inplace=True)

    results = []
    plot_files = []
    for test, group in df.groupby("Test"):
        group = group.sort_values("Expected_Conc")
        x = group["Expected_Conc"].values
        y = group["Measured_Conc"].values
        if len(x) < 3:
            continue
        coeffs = np.polyfit(x, y, 1)
        slope, intercept = coeffs
        y_pred = slope * x + intercept
        recovery = (y / x) * 100 if x[0] != 0 else np.nan  # avoid zero division
        # For 0 expected, recovery not defined; we can skip or set to NaN
        mask = x > 0
        avg_recovery = np.mean(recovery[mask])
        # Check recovery limits
        if test in ACCEPTANCE:
            low_rec, high_rec = ACCEPTANCE[test][3], ACCEPTANCE[test][4]
            pass_fail = status(low_rec <= avg_recovery <= high_rec)
            unit = ACCEPTANCE[test][0]
        else:
            low_rec, high_rec = None, None
            pass_fail = "Unknown"
            unit = "?"
        results.append({
            "Test": test,
            "Unit": unit,
            "Slope": round(slope, 3),
            "Intercept": round(intercept, 3),
            "Avg Recovery%": round(avg_recovery, 2) if not np.isnan(avg_recovery) else "N/A",
            "Recovery Range": f"{low_rec}-{high_rec}%" if low_rec else "N/A",
            "Status": pass_fail
        })

        # Plot
        plt.figure()
        plt.scatter(x, y, label="Measured")
        plt.plot(x, y_pred, 'r-', label=f"Fit: y={slope:.3f}x+{intercept:.3f}")
        plt.xlabel("Expected Concentration")
        plt.ylabel("Measured Concentration")
        plt.title(f"Linearity: {test}")
        plt.legend()
        fname = os.path.join(OUTPUT_DIR, f"linearity_{test}.png")
        plt.savefig(fname)
        plt.close()
        plot_files.append(fname)

    summary = pd.DataFrame(results)

    wb = Workbook()
    ws = wb.active
    ws.title = "Linearity Results"
    for r in dataframe_to_rows(summary, index=False, header=True):
        ws.append(r)
    img_row = len(summary) + 3
    for f in plot_files:
        plot_to_excel(ws, f, f"A{img_row}")
        img_row += 20
    out_path = os.path.join(OUTPUT_DIR, "Linearity_Report.xlsx")
    wb.save(out_path)
    print(f"Linearity report saved to {out_path}")
    return summary

# ------------------------------
# 5. CARRYOVER MODULE
# ------------------------------
def generate_carryover_template():
    """Template: Test, High1, High2, High3, Low1, Low2, Low3"""
    template = pd.DataFrame(columns=["Test", "High1", "High2", "High3", "Low1", "Low2", "Low3"])
    example = {"Test": "Glucose", "High1": "", "High2": "", "High3": "", "Low1": "", "Low2": "", "Low3": ""}
    template = pd.concat([template, pd.DataFrame([example])], ignore_index=True)
    path = os.path.join(TEMPLATE_DIR, "carryover_template.xlsx")
    template.to_excel(path, index=False)
    print(f"Carryover template saved to {path}")

def process_carryover(filename):
    df = pd.read_excel(filename)
    # Melt high and low into long format
    high_cols = ["High1", "High2", "High3"]
    low_cols = ["Low1", "Low2", "Low3"]
    results = []
    for _, row in df.iterrows():
        test = row["Test"]
        highs = pd.to_numeric(row[high_cols], errors='coerce').dropna()
        lows = pd.to_numeric(row[low_cols], errors='coerce').dropna()
        if len(highs) < 2 or len(lows) < 2:
            continue
        high_mean = highs.mean()
        low_mean = lows.mean()
        # Carryover% = (Low1 - Low3_avg) / (High2_avg - Low3_avg) * 100 (standard formula)
        # Using typical carryover formula: ((L1 - L3) / (H2 - L3)) * 100
        # We'll use mean of low replicates as L_avg, but better to follow CLSI: use the first low after high (Low1) and average of subsequent lows.
        # Simplification: Carryover = ((Low1 - mean of Low2,Low3) / (mean of High1-3 - mean of Low2,Low3)) * 100
        if len(lows) >= 3:
            low_first = lows.iloc[0]
            low_rest_mean = lows.iloc[1:].mean()
            high_all_mean = highs.mean()
            carry = ((low_first - low_rest_mean) / (high_all_mean - low_rest_mean)) * 100 if high_all_mean != low_rest_mean else np.nan
        else:
            carry = np.nan
        if test in ACCEPTANCE:
            limit = ACCEPTANCE[test][5]
            pass_fail = status(abs(carry) <= limit) if not np.isnan(carry) else "Invalid"
            unit = ACCEPTANCE[test][0]
        else:
            limit = None
            unit = "?"
            pass_fail = "Unknown"
        results.append({
            "Test": test,
            "Unit": unit,
            "Carryover%": round(carry, 2) if not np.isnan(carry) else "N/A",
            "Max Allowed%": limit,
            "Status": pass_fail
        })
    summary = pd.DataFrame(results)

    wb = Workbook()
    ws = wb.active
    ws.title = "Carryover Results"
    for r in dataframe_to_rows(summary, index=False, header=True):
        ws.append(r)
    # No plot, simple table
    out_path = os.path.join(OUTPUT_DIR, "Carryover_Report.xlsx")
    wb.save(out_path)
    print(f"Carryover report saved to {out_path}")
    return summary

# ------------------------------
# Command-line interface
# ------------------------------
def main():
    parser = argparse.ArgumentParser(description="BK-310 Biochemistry Validation")
    parser.add_argument("--module", type=str, required=True,
                        choices=["qc", "precision", "accuracy", "linearity", "carryover"],
                        help="Validation module to run")
    parser.add_argument("--generate", action="store_true", help="Generate empty template")
    parser.add_argument("--process", type=str, help="Process filled template file (xlsx path)")
    parser.add_argument("--replicates", type=int, default=10, help="Number of replicates for precision template")

    args = parser.parse_args()

    module = args.module
    if args.generate:
        if module == "qc":
            generate_qc_template()
        elif module == "precision":
            generate_precision_template(args.replicates)
        elif module == "accuracy":
            generate_accuracy_template()
        elif module == "linearity":
            generate_linearity_template()
        elif module == "carryover":
            generate_carryover_template()
    elif args.process:
        file_path = args.process
        if not os.path.exists(file_path):
            print(f"File {file_path} not found.")
            sys.exit(1)
        if module == "qc":
            process_qc(file_path)
        elif module == "precision":
            process_precision(file_path)
        elif module == "accuracy":
            process_accuracy(file_path)
        elif module == "linearity":
            process_linearity(file_path)
        elif module == "carryover":
            process_carryover(file_path)
    else:
        print("Please specify either --generate or --process <file>. Use --help for details.")

if __name__ == "__main__":
    main()
